"""
Script to generate a professional Excel report from raw_data.csv
Run: python scripts/generate_excel.py
"""

import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.chart.series import DataPoint

# ── Load & clean data ──────────────────────────────────────────────────────────
df = pd.read_csv("data/raw_data.csv")
df = df.drop_duplicates()
df["salary"] = df["salary"].fillna(df["salary"].mean())
df["performance_score"] = df["performance_score"].fillna(df["performance_score"].mean())
df["department"] = df["department"].fillna("Unknown")

os.makedirs("data", exist_ok=True)

wb = Workbook()

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 ─ Raw Data (styled)
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "📋 Dataset"

# Header style
HEADER_FILL   = PatternFill("solid", fgColor="4C4CFF")
HEADER_FONT   = Font(bold=True, color="FFFFFF", size=11, name="Segoe UI")
HEADER_ALIGN  = Alignment(horizontal="center", vertical="center")
ALT_FILL      = PatternFill("solid", fgColor="F0F0FF")
BORDER_SIDE   = Side(style="thin", color="CCCCCC")
CELL_BORDER   = Border(left=BORDER_SIDE, right=BORDER_SIDE,
                        top=BORDER_SIDE,  bottom=BORDER_SIDE)

headers = list(df.columns)
for col_idx, h in enumerate(headers, 1):
    cell = ws1.cell(row=1, column=col_idx, value=h.replace("_", " ").title())
    cell.fill      = HEADER_FILL
    cell.font      = HEADER_FONT
    cell.alignment = HEADER_ALIGN
    cell.border    = CELL_BORDER

for r_idx, row in enumerate(df.itertuples(index=False), 2):
    fill = PatternFill("solid", fgColor="F8F8FF") if r_idx % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
    for c_idx, val in enumerate(row, 1):
        cell = ws1.cell(row=r_idx, column=c_idx, value=val)
        cell.fill      = fill
        cell.border    = CELL_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font      = Font(name="Segoe UI", size=10)
        # Salary → currency format
        if headers[c_idx-1] == "salary":
            cell.number_format = "$#,##0"
        # Performance → percentage display
        if headers[c_idx-1] == "performance_score":
            cell.number_format = "0.0"

# Column widths
col_widths = {"age":10, "salary":14, "department":18,
              "years_experience":20, "performance_score":20, "city":14}
for col_idx, h in enumerate(headers, 1):
    ws1.column_dimensions[get_column_letter(col_idx)].width = col_widths.get(h, 14)

ws1.row_dimensions[1].height = 28
ws1.freeze_panes = "A2"

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 ─ Summary Statistics
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("📊 Summary")

TITLE_FONT  = Font(bold=True, size=14, color="4C4CFF", name="Segoe UI")
LABEL_FONT  = Font(bold=True, size=10, name="Segoe UI")
VALUE_FONT  = Font(size=10, name="Segoe UI")
LABEL_FILL  = PatternFill("solid", fgColor="EEF0FF")

ws2["B2"] = "📊 EDA Summary Report"
ws2["B2"].font      = Font(bold=True, size=16, color="4C4CFF", name="Segoe UI")
ws2["B2"].alignment = Alignment(horizontal="left")
ws2.merge_cells("B2:F2")

numeric_cols = df.select_dtypes(include="number").columns.tolist()
stats_df = df[numeric_cols].describe().round(2)

# Write stats header
for c_idx, col in enumerate(["Statistic"] + numeric_cols, 2):
    cell = ws2.cell(row=4, column=c_idx, value=col.replace("_", " ").title())
    cell.fill      = HEADER_FILL
    cell.font      = HEADER_FONT
    cell.alignment = HEADER_ALIGN
    cell.border    = CELL_BORDER

for r_idx, (stat_name, stat_row) in enumerate(stats_df.iterrows(), 5):
    fill = PatternFill("solid", fgColor="F8F8FF") if r_idx % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
    ws2.cell(row=r_idx, column=2, value=stat_name.capitalize()).fill = LABEL_FILL
    ws2.cell(row=r_idx, column=2).font = LABEL_FONT
    ws2.cell(row=r_idx, column=2).border = CELL_BORDER
    for c_idx, val in enumerate(stat_row, 3):
        cell = ws2.cell(row=r_idx, column=c_idx, value=round(float(val), 2))
        cell.fill   = fill
        cell.font   = VALUE_FONT
        cell.border = CELL_BORDER
        cell.alignment = Alignment(horizontal="center")

for col_idx in range(2, 2 + len(numeric_cols) + 1):
    ws2.column_dimensions[get_column_letter(col_idx)].width = 18
ws2.row_dimensions[4].height = 24

# ── Department Summary table (for chart) ──────────────────────────────────────
dept_summary = df.groupby("department").agg(
    Count=("age", "count"),
    Avg_Salary=("salary", "mean"),
    Avg_Score=("performance_score", "mean")
).round(0).reset_index()

start_row = stats_df.shape[0] + 7
ws2.cell(row=start_row, column=2, value="Department Summary").font = TITLE_FONT
ws2.merge_cells(f"B{start_row}:F{start_row}")

dept_heads = ["Department", "Count", "Avg Salary ($)", "Avg Performance"]
for c_idx, h in enumerate(dept_heads, 2):
    cell = ws2.cell(row=start_row+1, column=c_idx, value=h)
    cell.fill = HEADER_FILL; cell.font = HEADER_FONT
    cell.alignment = HEADER_ALIGN; cell.border = CELL_BORDER

for r_idx, row in enumerate(dept_summary.itertuples(index=False), start_row+2):
    fill = PatternFill("solid", fgColor="F8F8FF") if r_idx % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
    vals = [row.department, row.Count, row.Avg_Salary, row.Avg_Score]
    for c_idx, val in enumerate(vals, 2):
        cell = ws2.cell(row=r_idx, column=c_idx, value=val)
        cell.fill = fill; cell.border = CELL_BORDER
        cell.alignment = Alignment(horizontal="center")
        cell.font = VALUE_FONT

# ── Bar Chart: Dept vs Avg Salary ─────────────────────────────────────────────
chart_row      = start_row + 1
data_ref_start = chart_row + 1
data_ref_end   = data_ref_start + len(dept_summary) - 1

bar = BarChart()
bar.type   = "col"
bar.title  = "Average Salary by Department"
bar.y_axis.title = "Avg Salary ($)"
bar.x_axis.title = "Department"
bar.style  = 10
bar.width  = 20
bar.height = 12

data_ref = Reference(ws2, min_col=4, min_row=chart_row, max_row=data_ref_end)
cats_ref = Reference(ws2, min_col=2, min_row=data_ref_start, max_row=data_ref_end)
bar.add_data(data_ref, titles_from_data=True)
bar.set_categories(cats_ref)
ws2.add_chart(bar, f"H{chart_row}")

# ── Pie Chart: Dept headcount ─────────────────────────────────────────────────
pie = PieChart()
pie.title  = "Headcount by Department"
pie.style  = 10
pie.width  = 14
pie.height = 12

pie_data = Reference(ws2, min_col=3, min_row=chart_row, max_row=data_ref_end)
pie_cats = Reference(ws2, min_col=2, min_row=data_ref_start, max_row=data_ref_end)
pie.add_data(pie_data, titles_from_data=True)
pie.set_categories(pie_cats)
ws2.add_chart(pie, f"H{chart_row + 16}")

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 ─ Cleaned Data export
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("✅ Cleaned Data")
ws3["B2"] = "Cleaned & Deduplicated Dataset"
ws3["B2"].font = TITLE_FONT

for c_idx, h in enumerate(headers, 2):
    cell = ws3.cell(row=4, column=c_idx, value=h.replace("_", " ").title())
    cell.fill = HEADER_FILL; cell.font = HEADER_FONT
    cell.alignment = HEADER_ALIGN; cell.border = CELL_BORDER

for r_idx, row in enumerate(df.itertuples(index=False), 5):
    fill = PatternFill("solid", fgColor="F8F8FF") if r_idx % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
    for c_idx, val in enumerate(row, 2):
        cell = ws3.cell(row=r_idx, column=c_idx, value=val)
        cell.fill = fill; cell.border = CELL_BORDER
        cell.alignment = Alignment(horizontal="center"); cell.font = Font(name="Segoe UI", size=10)
        if headers[c_idx-2] == "salary":
            cell.number_format = "$#,##0"

for c_idx in range(2, len(headers) + 2):
    ws3.column_dimensions[get_column_letter(c_idx)].width = 18

# ── Save ───────────────────────────────────────────────────────────────────────
out_path = "data/EDA_Report_Mohamed_Mostafa.xlsx"
wb.save(out_path)
print(f"✅ Excel report saved → {out_path}")
print(f"   Sheets: {[s.title for s in wb.worksheets]}")
print(f"   Rows  : {len(df)} (after cleaning)")
